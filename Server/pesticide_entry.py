import json
from pprint import pprint
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import re


def read_accessories(file_path: str, starting_row: int = 4, end_col=6) -> str:
    possible_wb = load_workbook(file_path, read_only=True)

    if not possible_wb:
        print("Can't open workbook, check file")

    wb: Workbook = possible_wb

    ws: Worksheet = wb.active
    funcs = []
    for ws in wb.worksheets:
        # ws: Worksheet = wb[sheetname]

        funcs.append(ws.title)

        for row in ws.iter_rows(min_row=starting_row, max_col=end_col, values_only=True):
            # if row[-1] and str.isnumeric(row[-1]):

            name = str(row[end_col-3]).strip()
            dose_raw = str(row[end_col - 2]).strip()
            phi_raw = str(row[end_col - 1]).strip()

            # print(name)
            # print(dose_raw)
            # print(phi_raw)

            # if dose := re.match(r'(\d+,?\d*)', dose_raw):
            #     dose  = dose.groups()[0].replace(',', '.')

            if phi := re.match(r'(\d+)', phi_raw):
                phi = phi.groups()[0]
            else:
                phi = 0

            funcs.append(f"ShownPesticide('{name}', '{dose_raw}', {phi}),")

        funcs.append('\n')

    wb.close()
    return '\n'.join(funcs)


if __name__ == '__main__':
    filename = sys.argv[1]
    starting_row, end_col = 4, 6
    try:
        starting_row = int(sys.argv[2])
        end_col = int(sys.argv[3])
    except Exception as e:
        print(e)
        pass

    accessories = read_accessories(f'../pesticides/{filename}', starting_row,
                                   end_col)

    print(accessories)
